import cx_Oracle  
import pandas as pd
import json
from openpyxl import load_workbook
from datetime import datetime
from dateutil.parser import parse

def convert_data(item,result):
  
  lob_object = result[item]  
  lob_content = lob_object.read()
  lob_content_str = lob_content.encode('utf-8')
  lob_dict = str(json.loads(lob_content_str))
  return lob_dict

def format_time(item,result):
  formatted_date = result[item].strftime("%m/%d/%Y")
  return formatted_date
      
def export_data(data):
   
  workbook = load_workbook(filename='data.xlsx')
  sheet = workbook.active
  sheet.delete_rows(1, sheet.max_row)
  column_names = ['Tên khách hàng', 'Số điện thoại', 'Email','Loại khách hàng', 'CIF', 'CMND/CCCD', 'Ngày sinh', 'Nhóm đối tượng khách hàng', 'Địa chỉ thường trú', 'Địa chỉ liên hệ', 'Chức vụ', 'Ghi chú']
  sheet.append(column_names)
  for row in data:
    sheet.append(row)
  workbook.save(filename='data.xlsx')

def get_Cifs():
  df = pd.read_excel("input.xlsx")
  data = []
  for index, row in df.iterrows():
    data.append(row[4])
  my_string = '(' + ', '.join(data) + ')'
  cifs = my_string.replace('(', "('").replace(', ', "', '").replace(')', "')")

  return cifs
def main():
  try: 
    # Kết nối tới cơ sở dữ liệu Oracle
    dsn_tns = cx_Oracle.makedsn('10.0.18.116', '1521', service_name='orcl')
    connection = cx_Oracle.connect(user='los_test', password='123456', dsn=dsn_tns)

    cursor = connection.cursor()
    # Lấytime
    # day = (datetime.now()).strftime("%Y/%m/%d %H:%M:%S.%f")
    data = get_Cifs()
    
    query = "select cus_name,cus_phone,cus_email,retriction_name,cus_cif,cus_id,cus_dayofbirth,target_group_name,cus_permanent_address,cus_current_address,title_name,cus_note from los_2019.cf_lending_restriction where cus_cif in "+ data
    cursor.execute(query)

    rows = cursor.fetchall()
    data=[]

    # convert data
    for result in rows:
      result_copy = list(result)

      result_copy[6] = format_time(6,result)

      converted_result = tuple(result_copy)
      data.append((converted_result))

  # export file
    export_data(data)

  # Close
    cursor.close()
    connection.close()

    print("Successful")

  except cx_Oracle.DatabaseError as e: 
      print("There is a problem with Oracle", e) 

if __name__== "__main__":
  main()